#!/usr/bin/env python3
"""
Script pour chercher une valeur dans la colonne A d'un fichier Excel
et récupérer la valeur correspondante dans la colonne B.
"""

import sys
import argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌ La bibliothèque 'openpyxl' n'est pas installée.")
    print("📦 Installez-la avec: pip install openpyxl")
    sys.exit(1)


def search_in_excel(excel_file, search_value, column_search='A', columns_result=['B'], sheet_name=None):
    """
    Cherche une valeur dans une colonne et retourne les valeurs correspondantes dans d'autres colonnes.
    
    Args:
        excel_file (str): Chemin vers le fichier Excel
        search_value (str): Valeur à chercher
        column_search (str): Colonne où chercher (défaut: 'A')
        columns_result (list): Liste des colonnes dont on veut récupérer les valeurs (défaut: ['B'])
        sheet_name (str): Nom de la feuille à utiliser (None = feuille active)
    
    Returns:
        list: Liste de dictionnaires contenant les valeurs trouvées pour chaque ligne
    """
    # Ouvrir le fichier Excel
    workbook = openpyxl.load_workbook(excel_file)
    
    # Sélectionner la feuille
    if sheet_name:
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            workbook.close()
            raise ValueError(f"La feuille '{sheet_name}' n'existe pas. Feuilles disponibles: {', '.join(workbook.sheetnames)}")
    else:
        sheet = workbook.active
    
    results = []
    
    # Parcourir toutes les lignes
    for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row):
        # Récupérer la cellule de la colonne de recherche
        search_cell = None
        result_cells = {}
        
        for cell in row:
            if cell.column_letter == column_search:
                search_cell = cell
            if cell.column_letter in columns_result:
                result_cells[cell.column_letter] = cell
        
        # Vérifier si la valeur correspond
        if search_cell and search_cell.value is not None:
            # Conversion en string pour la comparaison
            if str(search_cell.value).strip() == str(search_value).strip():
                # Créer un dictionnaire avec toutes les colonnes demandées
                row_result = {'row': search_cell.row}
                for col in columns_result:
                    if col in result_cells and result_cells[col].value is not None:
                        row_result[col] = result_cells[col].value
                    else:
                        row_result[col] = None
                
                results.append(row_result)
    
    workbook.close()
    return results


def main():
    parser = argparse.ArgumentParser(
        description="Cherche une valeur dans la colonne A d'un fichier Excel et récupère la valeur en colonne B."
    )
    parser.add_argument(
        "-f", "--file",
        required=True,
        help="Chemin vers le fichier Excel (.xlsx)"
    )
    parser.add_argument(
        "-s", "--search",
        required=True,
        help="Valeur à rechercher"
    )
    parser.add_argument(
        "-c", "--column-search",
        default='A',
        help="Colonne où chercher (défaut: A)"
    )
    parser.add_argument(
        "-r", "--columns-result",
        default='B',
        help="Colonne(s) dont on veut récupérer les valeurs, séparées par des virgules (défaut: B). Exemple: B,C,D"
    )
    parser.add_argument(
        "--sheet",
        help="Nom de la feuille à utiliser (optionnel, utilise la feuille active par défaut)"
    )
    parser.add_argument(
        "--list-sheets",
        action="store_true",
        help="Afficher la liste des feuilles disponibles et quitter"
    )
    
    args = parser.parse_args()
    
    # Vérifier que le fichier existe
    excel_file = Path(args.file)
    if not excel_file.exists():
        print(f"❌ Le fichier '{args.file}' n'existe pas.")
        sys.exit(1)
    
    if not excel_file.suffix in ['.xlsx', '.xlsm']:
        print(f"❌ Le fichier doit être au format Excel (.xlsx ou .xlsm).")
        sys.exit(1)
    
    # Option pour lister les feuilles
    if args.list_sheets:
        try:
            workbook = openpyxl.load_workbook(excel_file)
            print(f"📋 Feuilles disponibles dans '{excel_file.name}':")
            for i, sheet_name in enumerate(workbook.sheetnames, 1):
                active_marker = " (active)" if workbook[sheet_name] == workbook.active else ""
                print(f"  {i}. {sheet_name}{active_marker}")
            workbook.close()
        except Exception as e:
            print(f"❌ Erreur lors de la lecture du fichier: {e}")
        sys.exit(0)
    
    # Parser les colonnes de résultat (séparées par des virgules)
    columns_result = [col.strip().upper() for col in args.columns_result.split(',')]
    
    print(f"📂 Fichier: {excel_file}")
    if args.sheet:
        print(f"📄 Feuille: {args.sheet}")
    print(f"🔍 Recherche de: '{args.search}' dans la colonne {args.column_search}")
    print(f"📊 Récupération des valeurs des colonnes: {', '.join(columns_result)}\n")
    
    # Effectuer la recherche
    try:
        results = search_in_excel(
            excel_file, 
            args.search,
            args.column_search,
            columns_result,
            args.sheet
        )
        
        if results:
            print(f"✅ {len(results)} résultat(s) trouvé(s):\n")
            for i, result in enumerate(results, 1):
                print(f"  Résultat {i} (ligne {result['row']}):")
                for col in columns_result:
                    value = result.get(col, 'N/A')
                    print(f"    - Colonne {col}: {value}")
                print()
        else:
            print(f"❌ Aucun résultat trouvé pour '{args.search}'")
    
    except Exception as e:
        print(f"❌ Erreur lors de la recherche: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
